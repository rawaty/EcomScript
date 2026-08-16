"""
Bulk Listing Generator (Meesho + Flipkart)
Streamlit UI — platform detection, template fill, QC, download.
"""

import base64
import io
import json
import os
import random
import re

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

from ai_helper import ai_suggest_fields, ai_analyze_image
from database import (
    get_supabase_client, check_duplicates, save_listings,
    get_listing_count, save_profile_cloud, load_profiles_cloud,
)
from excel_utils import (
    find_data_sheet, find_header_row, get_col_map,
    get_compulsory, get_dropdowns, inject_data,
)
from listing_builder import build_listing_rows, gen_sku
from platform_config import (
    SKIP_FIELDS,
    detect_platform, resolve_fields, programmatic_field_set, auto_field_set,
    scrub_restricted, detection_uncertain, restricted_keywords_for,
    PLATFORM_ALIASES,
)
from qc_checker import run_qc_check


PROFILES_FILE = "profiles.json"
PRESETS_FILE = "presets.json"

DEFAULT_PRESETS = {
    "Flip Flops (Men)": {
        "Generic Name": "Flip Flops", "Material": "EVA", "Sole Material": "EVA",
        "Type": "Thong Flip-flops", "Pattern": "Solid", "Waterproof": "Yes",
        "Main Trend": "Solid/Regular", "Net Quantity (N)": "1",
        "Net Weight (gms)": "220", "GST %": "5", "HSN ID": "64022090",
    },
    "Flip Flops (Women)": {
        "Generic Name": "Flip Flops", "Material": "PU", "Sole Material": "PU",
        "Type": "Thong Flip-flops", "Pattern": "Printed", "Waterproof": "No",
        "Main Trend": "Floral", "Net Quantity (N)": "1",
        "Net Weight (gms)": "180", "GST %": "5", "HSN ID": "64022090",
    },
    "Kurta Sets (Girls)": {
        "Generic Name": "Kurtis & Kurtas", "Occasion": "Casual",
        "Sleeve Length": "Long Sleeves", "Fabric": "Cotton",
        "Pattern": "Printed", "Net Weight (gms)": "300",
        "GST %": "5", "HSN ID": "6111", "Net Quantity (N)": "1",
        "Stitch Type": "Stitched",
    },
    "Sliders (Men)": {
        "Generic Name": "Sliders", "Material": "Synthetic", "Sole Material": "Rubber",
        "Type": "Sliders", "Pattern": "Solid", "Waterproof": "No",
        "Main Trend": "Solid/Regular", "Net Quantity (N)": "1",
        "Net Weight (gms)": "250", "GST %": "5", "HSN ID": "64022090",
    },
}


def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_json(filepath, data):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def parse_csv(raw):
    seen, result = set(), []
    for item in raw.split(","):
        t = item.strip()
        if t and t not in seen:
            seen.add(t)
            result.append(t)
    return result


parse_comma_separated = parse_csv


def parse_urls(raw):
    return [u.strip() for u in raw.strip().split("\n") if u.strip()]


def main():
    st.set_page_config(page_title="Bulk Listing Generator", layout="wide")
    st.title("🛒 Bulk Listing Generator (Meesho + Flipkart)")
    st.caption("Template upload → Dynamic form → Bulk fill → QC Check → Download")

    uploaded = st.file_uploader("📁 Upload your template (.xlsx)", type=["xlsx"])
    if not uploaded:
        st.info("👆 Upload a blank Meesho or Flipkart template to get started")
        return

    try:
        wb = load_workbook(uploaded, keep_vba=False)
        ws = find_data_sheet(wb)
        hdr = find_header_row(ws)
        col_map = get_col_map(ws, hdr)
        compulsory = get_compulsory(ws, hdr, col_map)
        dropdowns = get_dropdowns(wb, col_map)
    except Exception as e:
        st.error(f"Error reading template: {e}")
        return

    fields = [f for f in col_map if f not in SKIP_FIELDS]
    platform = detect_platform(col_map)
    resolved = resolve_fields(col_map, platform)
    auto_fields = auto_field_set(resolved)
    programmatic_fields = programmatic_field_set(resolved)
    variation_field = resolved.get("variation")
    desc_field = resolved.get("description")
    color_field = resolved.get("color")
    occasion_field = resolved.get("occasion")
    fabric_field = resolved.get("fabric")
    image_fields = [resolved[k] for k in ("image1", "image2", "image3", "image4") if k in resolved]

    if detection_uncertain(col_map):
        st.warning(
            "Could not confidently detect platform from headers — defaulting to Meesho. "
            "Flipkart aliases apply when Seller SKU / Selling Price headers are present."
        )

    st.success(
        f"✅ Platform: **{platform}** | Sheet: '{ws.title}' | {len(fields)} fields | "
        f"{len(compulsory)} required | {len(dropdowns)} dropdowns"
    )

    with st.expander("📋 All detected fields"):
        for f in fields:
            m = "⭐" if f in compulsory else "○"
            d = f" 🔽[{len(dropdowns[f])} opts]" if f in dropdowns else ""
            st.text(f"{m} {f}{d}")
        st.caption(f"Resolved aliases: {resolved}")

    st.markdown("---")
    _render_ai_image_fill(platform, dropdowns, fabric_field, occasion_field, desc_field)

    st.markdown("---")
    profiles, all_presets, selected_profile, selected_preset = _render_profiles_presets()
    _apply_prefills(profiles, all_presets, selected_profile, selected_preset, programmatic_fields)

    if st.button("🗑️ Clear prefill"):
        for k in [k for k in st.session_state if k.startswith("f_")]:
            del st.session_state[k]
        st.rerun()

    submitted, form_data = _render_main_form(
        dropdowns, color_field, variation_field, fields, compulsory,
        auto_fields, programmatic_fields, image_fields, resolved,
    )

    _render_save_profile()

    if not submitted:
        return

    _run_generation(
        form_data=form_data,
        platform=platform,
        resolved=resolved,
        col_map=col_map,
        compulsory=compulsory,
        dropdowns=dropdowns,
        programmatic_fields=programmatic_fields,
        uploaded=uploaded,
    )


def _render_ai_image_fill(platform, dropdowns, fabric_field, occasion_field, desc_field):
    st.markdown("### 📸 Smart Fill: Upload product photo → AI auto-fills form")
    product_image = st.file_uploader(
        "Upload product image (optional)", type=["jpg", "jpeg", "png"], key="product_img"
    )
    if not (product_image and os.environ.get("GEMINI_API_KEY")):
        return

    img_bytes = base64.b64encode(product_image.read()).decode()
    product_image.seek(0)

    if "ai_analysis_done" not in st.session_state:
        valid_options = {}
        for field in [
            "Color", "Fabric", "Pattern", "Occasion", "Sleeve Length",
            "Fit/ Shape", "Length", "Neck/Collar", "Print or Pattern Type",
            "Material",
        ]:
            if field in dropdowns:
                valid_options[field] = dropdowns[field][:30]

        with st.spinner("🤖 AI analyzing image..."):
            analysis = ai_analyze_image(
                img_bytes, category_hint="", valid_options=valid_options, platform=platform
            )

        if analysis:
            st.success("✅ AI detected product details!")
            st.json(analysis)
            if analysis.get("color"):
                st.session_state["ai_color"] = analysis["color"]
            if analysis.get("fabric"):
                st.session_state[f"f_{fabric_field or 'Fabric'}"] = analysis["fabric"]
            if analysis.get("pattern"):
                st.session_state["f_Pattern"] = analysis["pattern"]
            if analysis.get("occasion"):
                st.session_state[f"f_{occasion_field or 'Occasion'}"] = analysis["occasion"]
            if analysis.get("category"):
                st.session_state["ai_category"] = analysis["category"]
            if analysis.get("description"):
                desc_text = scrub_restricted(analysis["description"], platform)
                st.session_state["ai_description"] = desc_text
                if desc_field:
                    st.session_state[f"f_{desc_field}"] = desc_text
            if analysis.get("title_suggestion"):
                st.session_state["ai_title"] = analysis["title_suggestion"]
            st.session_state["ai_analysis_done"] = True
            st.rerun()
        else:
            st.warning("AI analysis failed — fill manually")
        return

    st.success("✅ AI auto-filled form fields below!")
    fabric_key = fabric_field or "Fabric"
    occ_key = occasion_field or "Occasion"
    ai_data = {}
    if st.session_state.get("ai_color"):
        ai_data["color"] = st.session_state["ai_color"]
    if st.session_state.get(f"f_{fabric_key}"):
        ai_data["fabric"] = st.session_state[f"f_{fabric_key}"]
    if st.session_state.get("f_Pattern"):
        ai_data["pattern"] = st.session_state["f_Pattern"]
    if st.session_state.get(f"f_{occ_key}"):
        ai_data["occasion"] = st.session_state[f"f_{occ_key}"]
    if st.session_state.get("ai_category"):
        ai_data["category"] = st.session_state["ai_category"]
    if st.session_state.get("ai_title"):
        ai_data["title_suggestion"] = st.session_state["ai_title"]
    if st.session_state.get("ai_description"):
        ai_data["description"] = st.session_state["ai_description"]
    with st.expander("🤖 AI Analysis Result (JSON)", expanded=True):
        st.json(ai_data)
    if st.button("🔄 Re-analyze image"):
        del st.session_state["ai_analysis_done"]
        st.rerun()


def _render_profiles_presets():
    prof_col, preset_col = st.columns(2)
    with prof_col:
        st.markdown("**👤 Saved Profiles**")
        db_client_profiles = get_supabase_client()
        profiles = (
            load_profiles_cloud(db_client_profiles)
            if db_client_profiles else load_json(PROFILES_FILE)
        )
        profile_names = list(profiles.keys())
        selected_profile = (
            st.selectbox("Load Profile", ["-- None --"] + profile_names, key="profile_select")
            if profile_names else "-- None --"
        )
    with preset_col:
        st.markdown("**📦 Category Presets**")
        all_presets = {**DEFAULT_PRESETS, **load_json(PRESETS_FILE)}
        selected_preset = st.selectbox(
            "Load Preset", ["-- None --"] + list(all_presets.keys()), key="preset_select"
        )
    return profiles, all_presets, selected_profile, selected_preset


def _apply_prefills(profiles, all_presets, selected_profile, selected_preset, programmatic_fields):
    if selected_profile != "-- None --" and selected_profile in profiles:
        for k, v in profiles[selected_profile].items():
            if k not in programmatic_fields:
                st.session_state[f"f_{k}"] = v
    if selected_preset != "-- None --" and selected_preset in all_presets:
        for k, v in all_presets[selected_preset].items():
            if k not in programmatic_fields:
                st.session_state[f"f_{k}"] = v
    for stale in [f"f_{f}" for f in programmatic_fields]:
        st.session_state.pop(stale, None)


def _render_main_form(
    dropdowns, color_field, variation_field, fields, compulsory,
    auto_fields, programmatic_fields, image_fields, resolved,
):
    with st.form("main_form"):
        st.markdown("### 🎨 Core Settings")
        c1, c2 = st.columns(2)
        with c1:
            color_options = dropdowns.get(color_field, []) if color_field else []
            if color_options:
                default_colors = []
                ai_color = st.session_state.get("ai_color", "")
                if ai_color and ai_color in color_options:
                    default_colors = [ai_color]
                colors_raw = ", ".join(
                    st.multiselect("Colors *", options=color_options, default=default_colors)
                )
            else:
                colors_raw = st.text_input(
                    "Colors *",
                    value=st.session_state.get("ai_color", ""),
                    placeholder="Black, Blue, Red",
                )
            brand = st.text_input("Brand Name *", placeholder="kidoready")
            category = st.text_input(
                "Product Category *",
                value=st.session_state.get("ai_category", ""),
                placeholder="Kurtis & Kurtas",
            )
        with c2:
            size_options = dropdowns.get(variation_field, []) if variation_field else []
            if size_options:
                sizes_raw = ", ".join(st.multiselect("Sizes *", options=size_options))
            else:
                sizes_raw = st.text_input("Sizes *", placeholder="3-4 Years, 5-6 Years")
            style_code = st.text_input("Base Style Code *", placeholder="Kurti-01")
            audience = st.selectbox(
                "Target Audience *",
                [
                    "for Boys", "for Girls", "for Kids", "for Baby Boys",
                    "for Baby Girls", "for Men", "for Women", "Unisex",
                ],
            )

        st.markdown("### 🤖 AI & Database")
        ai_c1, ai_c2 = st.columns(2)
        with ai_c1:
            use_ai_titles = st.checkbox("✨ AI Titles (Gemini)")
            use_ai_desc = st.checkbox("📝 AI Descriptions")
        with ai_c2:
            use_ai_suggest = st.checkbox("💡 AI Field Suggestions")
            gemini_key = st.text_input(
                "Gemini Key", value=os.environ.get("GEMINI_API_KEY", ""), type="password"
            )

        st.markdown("### 💰 Price & Count")
        c3, c4, c5 = st.columns(3)
        with c3:
            count = st.number_input("Listings count", 1, 5000, 50, 10)
        with c4:
            price_var = st.number_input("Price variation (±₹)", 0, 100, 0, 5)
        with c5:
            mrp_val = st.text_input("MRP", placeholder="599")

        st.markdown("### 🖼️ Images (one URL per line)")
        ic1, ic2 = st.columns(2)
        with ic1:
            img1_raw = st.text_area(
                resolved.get("image1", "Image 1") + " *",
                height=80,
                placeholder="https://example.com/front.jpg",
            )
            img2_raw = st.text_area(resolved.get("image2", "Image 2"), height=80)
        with ic2:
            img3_raw = st.text_area(resolved.get("image3", "Image 3"), height=80)
            img4_raw = st.text_area(resolved.get("image4", "Image 4"), height=80)

        st.markdown("### 📝 Template Fields")
        fv = {}
        skip_form = set(auto_fields) | set(programmatic_fields) | set(image_fields)
        if variation_field:
            skip_form.add(variation_field)
        skip_form.update({"Foot Length Size", "Foot Width Size"})
        show = [f for f in fields if f not in skip_form]
        req = [f for f in show if f in compulsory]
        opt = [f for f in show if f not in compulsory]

        if req:
            st.markdown("**⭐ Required:**")
            cols = st.columns(2)
            for i, f in enumerate(req):
                with cols[i % 2]:
                    if f in dropdowns:
                        fv[f] = st.selectbox(f"⭐ {f}", [""] + dropdowns[f], key=f"f_{f}")
                    else:
                        fv[f] = st.text_input(f"⭐ {f}", key=f"f_{f}")
        if opt:
            with st.expander(f"📎 Optional Fields ({len(opt)})"):
                cols2 = st.columns(2)
                for i, f in enumerate(opt):
                    with cols2[i % 2]:
                        if f in dropdowns:
                            fv[f] = st.selectbox(f, [""] + dropdowns[f], key=f"f_{f}")
                        else:
                            fv[f] = st.text_input(f, key=f"f_{f}")

        submitted = st.form_submit_button("🚀 Generate & Fill Template")

    form_data = {
        "colors_raw": colors_raw,
        "sizes_raw": sizes_raw,
        "brand": brand,
        "category": category,
        "style_code": style_code,
        "audience": audience,
        "use_ai_titles": use_ai_titles,
        "use_ai_desc": use_ai_desc,
        "use_ai_suggest": use_ai_suggest,
        "gemini_key": gemini_key,
        "count": count,
        "price_var": price_var,
        "mrp_val": mrp_val,
        "img1_raw": img1_raw,
        "img2_raw": img2_raw,
        "img3_raw": img3_raw,
        "img4_raw": img4_raw,
        "fv": fv,
    }
    return submitted, form_data


def _render_save_profile():
    with st.expander("💾 Save / Update Profile"):
        save_col1, save_col2, save_col3 = st.columns([3, 1, 1])
        with save_col1:
            profile_name = st.text_input("Profile name", placeholder="My Business")
        with save_col2:
            st.markdown("")
            save_btn = st.button("💾 Save")
        with save_col3:
            st.markdown("")
            update_btn = st.button("🔄 Update")

        if not (save_btn or update_btn):
            return
        if not profile_name.strip():
            st.error("Profile name daalo")
            return

        skip_in_profile = {
            "Generic Name", "Color", "Pattern", "Fabric", "Occasion",
            "Fit/ Shape", "Length", "Neck/Collar", "Ornamentation",
            "Print or Pattern Type", "Sleeve Length",
            "Bottom Length Size", "Bottom Waist Size",
            "Duppatta Length Size", "Top Bust Size", "Top Length Size",
            "Material",
        }
        profile_data = {}
        for key, val in st.session_state.items():
            if key.startswith("f_"):
                field_name = key[2:]
                if field_name not in skip_in_profile and str(val).strip():
                    profile_data[field_name] = str(val).strip()

        if not profile_data:
            st.error("Koi field filled nahi hai — pehle form bharo fir save karo")
            return

        db_save = get_supabase_client()
        name = profile_name.strip()
        if db_save and save_profile_cloud(db_save, name, profile_data):
            action = "Updated" if update_btn else "Saved"
            st.success(f"✅ Profile '{name}' {action}! ({len(profile_data)} fields)")
            return

        profiles_local = load_json(PROFILES_FILE)
        profiles_local[name] = profile_data
        save_json(PROFILES_FILE, profiles_local)
        if db_save:
            st.warning("⚠️ Cloud failed — saved locally")
        else:
            st.success(f"✅ Profile '{name}' saved locally!")


def _run_generation(
    *, form_data, platform, resolved, col_map, compulsory, dropdowns,
    programmatic_fields, uploaded,
):
    colors_raw = form_data["colors_raw"]
    sizes_raw = form_data["sizes_raw"]
    brand = form_data["brand"]
    category = form_data["category"]
    style_code = form_data["style_code"]
    audience = form_data["audience"]
    fv = form_data["fv"]
    gemini_key = form_data["gemini_key"]

    errs = []
    if not colors_raw.strip():
        errs.append("Colors required")
    if not sizes_raw.strip():
        errs.append("Sizes required")
    if not brand.strip():
        errs.append("Brand Name required")
    if not style_code.strip():
        errs.append("Style Code required")
    if not category.strip():
        errs.append("Product Category required")
    for f in fv:
        if f in compulsory and not str(fv[f]).strip():
            errs.append(f"⭐ '{f}' is required")
    if errs:
        for e in errs:
            st.error(e)
        return

    colors = parse_csv(colors_raw)
    sizes = parse_csv(sizes_raw)

    if form_data["use_ai_suggest"] and gemini_key:
        os.environ["GEMINI_API_KEY"] = gemini_key
        suggestions = ai_suggest_fields(
            category.strip(), list(col_map.keys()), platform=platform
        )
        if suggestions:
            for field, value in suggestions.items():
                if field in fv and not str(fv[field]).strip():
                    fv[field] = value
            st.info(f"🤖 AI suggested {len(suggestions)} fields")

    img_url_lists = [
        parse_urls(form_data["img1_raw"]),
        parse_urls(form_data["img2_raw"]),
        parse_urls(form_data["img3_raw"]),
        parse_urls(form_data["img4_raw"]),
    ]

    # Persist AI descriptions in session + local cache
    desc_cache = {
        k: v for k, v in st.session_state.items() if isinstance(k, str) and k.startswith("desc_")
    }

    rows, actual_per_catalog, max_per_catalog = build_listing_rows(
        platform=platform,
        resolved=resolved,
        col_map=col_map,
        programmatic_fields=programmatic_fields,
        brand=brand,
        style_code=style_code,
        category=category,
        colors=colors,
        sizes=sizes,
        audience=audience,
        fv=fv,
        count=form_data["count"],
        price_var=form_data["price_var"],
        mrp_val=form_data["mrp_val"],
        img_url_lists=img_url_lists,
        use_ai_titles=form_data["use_ai_titles"],
        use_ai_desc=form_data["use_ai_desc"],
        gemini_key=gemini_key,
        desc_cache=desc_cache,
    )
    for k, v in desc_cache.items():
        st.session_state[k] = v

    if actual_per_catalog < int(form_data["count"]):
        st.warning(
            f"⚠️ Max per catalog = {len(colors)}×{len(sizes)} = {max_per_catalog}"
        )

    qc_errors, qc_warnings = run_qc_check(
        rows, col_map, dropdowns, platform=platform, fields=resolved
    )
    if qc_errors:
        st.error(f"❌ QC FAILED — {len(qc_errors)} error(s):")
        for err in qc_errors[:10]:
            st.markdown(f"  🔴 Row {err['row']} | `{err['field']}` — {err['message']}")
    if qc_warnings:
        with st.expander(f"⚠️ {len(qc_warnings)} Warning(s)"):
            for w in qc_warnings[:10]:
                st.markdown(f"  🟡 Row {w['row']} | `{w['field']}` — {w['message']}")
    if not qc_errors:
        st.success("✅ QC Pre-Check PASSED!")

    f_style = resolved.get("style_id")
    f_sku = resolved.get("sku")
    img_keys = [resolved.get(k) for k in ("image1", "image2", "image3", "image4")]

    db_client = None
    if os.environ.get("SUPABASE_URL") and os.environ.get("SUPABASE_KEY"):
        db_client = get_supabase_client()

    if db_client:
        all_style_ids = {r.get(f_style, "") for r in rows if f_style} - {""}
        all_sku_ids = {r.get(f_sku, "") for r in rows if f_sku} - {""}
        all_img_urls = set()
        for r in rows:
            for img_f in img_keys:
                if img_f and r.get(img_f):
                    all_img_urls.add(r[img_f])

        dupes = check_duplicates(
            db_client, all_style_ids, all_sku_ids, list(all_img_urls)[:20]
        )
        if dupes["sku_ids"] and f_sku:
            dupe_skus = set(dupes["sku_ids"])
            for row in rows:
                while row.get(f_sku, "") in dupe_skus:
                    row[f_sku] = gen_sku(style_code.strip(), random.randint(100, 9999))
            st.info(f"🔄 {len(dupe_skus)} duplicate SKU IDs detected — auto-regenerated!")
        if dupes["style_ids"]:
            st.error(f"🔴 DUPLICATE Style IDs! Already uploaded: {dupes['style_ids'][:3]}")
        if dupes["image_urls"]:
            st.warning(
                f"⚠️ {len(dupes['image_urls'])} image(s) pehle use hui — duplicate risk!"
            )
        st.caption(f"🗄️ Database: {get_listing_count(db_client)} listings stored")

    uploaded.seek(0)
    wb2 = load_workbook(uploaded, keep_vba=False)
    inject_data(wb2, rows)
    out = io.BytesIO()
    wb2.save(out)
    out.seek(0)
    safe_style = re.sub(r"[^A-Za-z0-9_-]+", "-", style_code.strip()) or "listing"
    filename = f"{platform}_{safe_style}_{len(rows)}.xlsx"
    st.success(f"✅ {len(rows)} listings generated for {platform}!")
    st.download_button(
        "📥 Download Filled Template",
        out.getvalue(),
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    df = pd.DataFrame(rows)
    prev_cols = [
        c for c in [
            resolved.get("product_name"), resolved.get("variation"),
            resolved.get("style_id"), resolved.get("group_id"),
            resolved.get("color"), resolved.get("sku"), resolved.get("price"),
        ]
        if c and c in df.columns
    ]
    st.dataframe(df[prev_cols].head(20) if prev_cols else df.head(20), use_container_width=True)

    if db_client and not qc_errors:
        if save_listings(db_client, rows, catalog_name=category.strip(), fields=resolved):
            st.success("💾 Listings saved to database!")


if __name__ == "__main__":
    main()
